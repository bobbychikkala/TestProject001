import openpyxl
import os



class Rd_excel:

    def __init__(self, file, sheetName):
        self.file = file
        self.workbook = openpyxl.load_workbook(self.file)
        self.sheet = self.workbook[sheetName]

    def getRowCont(self):
        self.rows = self.sheet.max_row
        return self.rows

    def getColumnCount(self):
        self.columns = self.sheet.max_column
        return self.columns

    def readData(self, rowNo, columnNo):
        data = self.sheet.cell(rowNo, columnNo).value
        return data

    def writeData(self, rowNo, columnNo, data):
        self.sheet.cell(rowNo, columnNo).value = data

        self.workbook.save(self.file)

# file = os.path.abspath(os.curdir)+ "\\testData\\testData.xlsx"  #"D:\\pythonWS\\Project_001\\testData\\testData.xlsx"
# sheet= "Sheet1"

# xlObj = Rd_excel(file,sheet)

# print(xlObj.getRowCont())
# print(xlObj.getColumnCount())

# print(xlObj.readData(1,2))
# xlObj.writeData(8,2,"SivaKumar")

# print(xlObj.readData(2,2))
# print(xlObj.readData(3,2))
# print(xlObj.readData(4,2))
